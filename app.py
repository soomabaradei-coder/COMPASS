"""COMPASS — نظام الإرشاد الأكاديمي الذكي (MVP).

قسم علوم الحاسب — كلية الحاسبات وتقنية المعلومات، جامعة الملك عبدالعزيز.
القاعدة الذهبية: كل رغبة تسجيل لا تُعتمد إلا من المرشدة الأكاديمية.
"""
import io
import os
import functools
from flask import (
    Flask, request, redirect, url_for, session, render_template, flash,
    abort, send_file,
)
from werkzeug.security import check_password_hash
from db import get_db, init_db
from plan_data import (BLOCKS, CHALLENGE_BLOCKS, CHALLENGE_TYPES, blocks_group_for_level)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "compass-kau-dev-secret-change-in-production")
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0  # لا تخزين مؤقت للملفات الثابتة أثناء التطوير

TYPE_AR = {"core": "Core", "elective": "Elective", "free": "Free"}
PARITY_AR = {"odd": "Odd", "even": "Even", "both": "Both"}
STATUS_AR = {"pending": "Pending", "approved": "Approved", "rejected": "Rejected"}

app.jinja_env.globals.update(TYPE_AR=TYPE_AR, PARITY_AR=PARITY_AR, STATUS_AR=STATUS_AR)


# ------------------------- مهام الطالبة (الشريط الجانبي) -------------------------
# رحلة الطالبة من التسجيل حتى التخرج. المتاح منها يعمل، والباقي "قريباً".
STUDENT_TASKS = [
    {"section": "Start Your Journey", "items": [
        {"key": "profile", "label": "Complete Profile", "icon": "",
         "endpoint": "student_profile", "status": "active",
         "desc": "Your academic details and advisor."},
    ]},
    {"section": "Every Semester", "items": [
        {"key": "wishes", "label": "Set Course Preferences", "icon": "",
         "endpoint": "student_wishes", "status": "active",
         "desc": "Select next semester's courses for advisor approval."},
        {"key": "blocks", "label": "Choose Block", "icon": "",
         "endpoint": "student_blocks", "status": "active",
         "desc": "Pick the ready-made schedule block for your level."},
        {"key": "schedule_waiver", "label": "Resolve Schedule Issues / Submit Waiver", "icon": "",
         "endpoint": "student_schedule_issues", "status": "active",
         "desc": "Report a registration problem, or submit a general waiver request."},
    ]},
    {"section": "Your Summer Experience", "items": [
        {"key": "summer_training", "label": "Apply for Summer Training Approval", "icon": "",
         "endpoint": "student_summer_training", "status": "active",
         "desc": "Request CPCS-323 (200 hrs) after completing level 7 & 8."},
    ]},
    {"section": "Your Final Year", "items": [
        {"key": "senior_project", "label": "Apply for Senior Project Approval", "icon": "",
         "endpoint": "student_senior_project", "status": "active",
         "desc": "Check registration eligibility for CPCS-498."},
        {"key": "sp_team", "label": "Register Your Senior Project Team", "icon": "",
         "endpoint": "student_sp_team", "status": "active",
         "desc": "Form your graduation-project team (3 members) after approval."},
        {"key": "graduation", "label": "Apply for Graduation Approval", "icon": "",
         "endpoint": "student_graduation", "status": "active",
         "desc": "Confirm you meet graduation requirements and request approval."},
    ]},
    {"section": "Before Leaving the University", "items": [
        {"key": "track_cert", "label": "Request Track Certificate", "icon": "",
         "endpoint": "student_track_cert", "status": "active",
         "desc": "Verify track-certificate eligibility against the plan."},
        {"key": "grad_letter", "label": "Download Graduation Letter", "icon": "",
         "endpoint": "student_grad_letter", "status": "active",
         "desc": "Download the plan-conformity statement after approval."},
    ]},
]

# فهرسة سريعة بالمفتاح ومطابقة نقطة النهاية بالمفتاح النشط
TASK_BY_KEY = {it["key"]: it for sec in STUDENT_TASKS for it in sec["items"]}
ENDPOINT_TO_TASK = {it["endpoint"]: it["key"] for sec in STUDENT_TASKS
                    for it in sec["items"] if it["endpoint"]}


# متطلبات التسجيل في مشروع التخرج CPCS-498 (من نموذج القسم)
SENIOR_PROJECT_PREREQS = [
    ("CPCS-223", "Analysis & Design of Algorithms"),
    ("CPIS-334", "Software Project Management"),
    ("CPCS-351", "Software Engineering I"),
    ("CPCS-241", "Database I"),
    ("CPCS-361", "Operating Systems I"),
    ("CPCS-331", "Artificial Intelligence I"),
]

# المسارات الستة ومقرراتها (من نموذج تجميع شهادات المسارات)
TRACKS = [
    {"key": "acp", "name": "Advanced Computer Programming",
     "courses": ["CPCS 403: Internet Application Programming",
                 "CPCS 466: System Programming (Competitive Programming I)",
                 "CPCS 494: Selected Topics (Competitive Programming II)",
                 "CPCS 405: Software Technology Topics"]},
    {"key": "ai", "name": "Intelligent Systems (Artificial Intelligence)",
     "courses": ["CPCS 432: Artificial Intelligence",
                 "CPCS 433: Artificial Intelligence Topics",
                 "CPCS 482: MultiMedia & User Interface Design",
                 "CPCS 494: Selected Topics (Big Data Analytics)"]},
    {"key": "se", "name": "Software Engineering",
     "courses": ["CPCS 353: Software Engineering Practices",
                 "CPCS 404: Component-Based Computing",
                 "CPCS 454: Object-Oriented Analysis and Design",
                 "CPCS 457: Software Engineering Theory"]},
    {"key": "hpc", "name": "High Performance Computing",
     "courses": ["CPCS 413: Computer Architecture II",
                 "CPCS 414: High Performance Computing",
                 "CPCS 494: Selected Topics (Cloud Computing / Big Data)"]},
    {"key": "net", "name": "Network Computing",
     "courses": ["CPCS 372: Computer Network 2",
                 "CPCS 473: Computer Network Practice",
                 "CPCS 474: TCP/IP & Web Networking"]},
    {"key": "sec", "name": "Information Security (Cyber Security)",
     "courses": ["CPCS 425: Information Security",
                 "CPCS 463: Computing Systems Security",
                 "CPCS 464: Dependable Computing",
                 "CPCS 494: Selected Topics (Computer Forensics)"]},
]
TRACK_BY_KEY = {t["key"]: t for t in TRACKS}

GRAD_SEMESTERS = ["First Semester", "Second Semester", "Summer Semester"]


@app.context_processor
def inject_student_nav():
    active = ENDPOINT_TO_TASK.get(request.endpoint)
    if request.endpoint == "student_task":
        active = request.view_args.get("key")
    if request.endpoint == "student_override":  # waiver shares the merged sidebar item
        active = "schedule_waiver"
    return {"student_tasks": STUDENT_TASKS, "active_task": active}


# ------------------------- المصادقة والصلاحيات -------------------------
def current_user():
    uid = session.get("uid")
    if not uid:
        return None
    return get_db().execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()


@app.context_processor
def inject_user():
    return {"user": current_user()}


def login_required(role=None):
    def deco(view):
        @functools.wraps(view)
        def wrapped(*args, **kwargs):
            u = current_user()
            if not u:
                return redirect(url_for("login"))
            if role and u["role"] != role:
                abort(403)
            return view(*args, **kwargs)
        return wrapped
    return deco


def get_student(user_id):
    return get_db().execute("SELECT * FROM students WHERE user_id=?", (user_id,)).fetchone()


# ------------------------- الدخول والخروج -------------------------
@app.route("/")
def index():
    u = current_user()
    if not u:
        return redirect(url_for("login"))
    return redirect(url_for(f"{u['role']}_home"))


@app.route("/about")
def about():
    return render_template("about.html")


@app.route("/sw.js")
def service_worker():
    # Served from root so the service worker's scope covers the whole app.
    resp = send_file(os.path.join(app.static_folder, "sw.js"),
                     mimetype="application/javascript")
    resp.headers["Cache-Control"] = "no-cache"
    resp.headers["Service-Worker-Allowed"] = "/"
    return resp


# ------------------------- COMPASS-Bot (answers from the CS plan only) -------------------------
import re as _re

_BOT_STOP = {"the", "a", "an", "is", "are", "what", "which", "how", "many", "of", "for",
             "to", "in", "do", "does", "i", "need", "course", "courses", "and", "or", "me",
             "my", "can", "take", "with", "prerequisite", "prerequisites", "prereq", "prereqs",
             "credit", "credits", "hours", "level", "plan", "about", "tell", "give", "list",
             "study", "cs", "you", "it", "this", "that", "have", "has"}


# Task guidance — what each task is and what it requires (grounded in the system's real rules)
TASK_HELP = [
    {"keys": ["course preference", "register course", "how to register", "registration preference",
              "course wishes", "how do i pick courses", "select courses", "how do i choose courses"],
     "text": "Course Preferences: during the open collection period, pick the courses you'd like to "
             "register next semester and submit them. They go to your academic advisor, who approves or "
             "rejects each one — nothing is registered without approval."},
    {"keys": ["choose block", "choosing a block", "pick a block", "block selection", "what is a block",
              "which block", "how to choose block", "how do i pick a block"],
     "text": "Choose Block: select the ready-made schedule block for your level (each block is a "
             "conflict-free set of courses). Your choice is sent to your advisor for approval."},
    {"keys": ["schedule issue", "registration problem", "report a problem", "schedule problem",
              "registration challenge", "couldn't add", "couldn't register", "cant add"],
     "text": "Schedule Issues: if you couldn't complete registration, pick the problem type, choose your "
             "level/block, and mark the affected sections (by their reference numbers). It's sent to your "
             "advisor to resolve."},
    {"keys": ["waiver", "override request", "exception request", "how to request waiver", "how do i waive"],
     "text": "Waiver Request: submit a general exception request. Optionally choose the related course, "
             "write the reason clearly, and send it. Your advisor reviews it and takes the needed action."},
    {"keys": ["how do i graduate", "graduation confirmation", "request graduation", "confirm graduation",
              "graduation process", "apply to graduate", "how to graduate"],
     "text": "Graduation Requirements: the system checks your record against 121 core credits, 9 elective "
             "courses, and 10 free credits (140 total). When ready, choose your graduation semester and "
             "submit a confirmation request; your advisor verifies and approves it, then the Plan "
             "Conformity Statement is issued."},
    {"keys": ["senior project team", "graduation project team", "form a team", "form my team",
              "project team", "team for graduation project", "how do i form my team"],
     "text": "Senior Project Team: after your Graduation Project (CPCS-498) eligibility is approved, form "
             "a group of 3 students (the third member is optional). Enter each member's name and KAU ID, "
             "choose your project supervisor, confirm you contacted her, then submit for approval."},
    {"keys": ["graduation project", "senior project", "project eligibility", "cpcs-498 eligib",
              "apply for graduation project", "cpcs 498 eligib"],
     "text": "Graduation Project (CPCS-498): you're eligible once you've passed CPCS-223, CPIS-334, "
             "CPCS-351, CPCS-241, CPCS-361, and CPCS-331. Mark the ones you've passed and submit; your "
             "advisor verifies and approves."},
    {"keys": ["summer training", "apply for summer training", "summer training requirement",
              "cpcs-323 requirement", "cpcs 323 requirement", "how do i do summer training"],
     "text": "Summer Training (CPCS-323, 200 hours): you must complete all level 7 & 8 courses. If you're "
             "missing one or two, you can apply as a waiver (department approval); missing more than two "
             "means you can't apply yet."},
    {"keys": ["track certificate", "track cert", "declare track", "my track", "how is my track",
              "determine my track", "choose track", "how do i get my track"],
     "text": "Track Certificate: your track is determined by your department elective courses. Choose your "
             "track and the two track courses you've completed, then submit. Your advisor approves it, and "
             "this sets your track."},
    {"keys": ["graduation letter", "plan conformity", "conformity statement", "get my letter", "my letter"],
     "text": "Graduation Letter (Plan Conformity Statement): issued after your graduation confirmation "
             "request is approved by your advisor. You can then view and print it."},
]


def bot_answer(q):
    """Answer strictly from the CS study plan and the system's task rules."""
    db = get_db()
    ql = (q or "").lower().strip()
    if not ql:
        return "Ask me about the CS study plan — a course, its prerequisites, credits, or level."

    # Info the plan does NOT contain — decline rather than return a non-answer
    _unsupported = ["who teach", "teacher", "instructor", "professor", "lecturer",
                    "what time", "meeting time", "class time", " room", "gpa", "pass rate",
                    "seats", "capacity", "how hard", "is it hard", "is it easy", "difficult",
                    " hard", " easy", "grade", "grades", "textbook", "office hour"]
    if any(u in ql for u in _unsupported):
        return ("The CS plan doesn't include that (e.g. instructors, timings, rooms, or grades), "
                "so I won't guess. I can answer courses, prerequisites, levels, blocks, or "
                "graduation requirements.")

    # task guidance — help the student understand what each task requires
    for entry in TASK_HELP:
        if any(k in ql for k in entry["keys"]):
            return entry["text"]

    # explicit course codes: "cpcs 204", "cpcs-204", "cpcs204"
    codes = [f"{a.upper()}-{b}" for a, b in _re.findall(r"\b([a-zA-Z]{3,4})[\s-]?(\d{3})\b", q)]
    if "summer training" in ql or "summer" in ql:
        codes.append("CPCS-323")
    if "graduation project" in ql or "senior project" in ql or "grad project" in ql:
        codes += ["CPCS-498", "CPCS-499"]

    found, seen = [], set()
    for code in codes:
        if code in seen:
            continue
        seen.add(code)
        r = db.execute("SELECT * FROM courses WHERE code=?", (code,)).fetchone()
        if r:
            found.append(r)

    wants_prereq = any(w in ql for w in ["prereq", "prerequisit", "before", "require"])
    if found:
        parts = []
        for r in found:
            if wants_prereq:
                parts.append(f"{r['code']} ({r['name']}) — prerequisite: {r['prereq'] or 'none'}.")
            else:
                pr = f" Prerequisite: {r['prereq']}." if r["prereq"] else ""
                parts.append(f"{r['code']} — {r['name']}: {r['credits']} credits, "
                             f"{TYPE_AR.get(r['type'], r['type']).lower()}, level {r['recommended_level']}.{pr}")
        return " ".join(parts)

    # graduation-requirement facts (from settings)
    if ("block" not in ql and
            any(w in ql for w in ["graduat", "how many credit", "total credit", "requirement", "graduate"])):
        reqs = {r["key"]: int(r["value"]) for r in db.execute("SELECT key,value FROM settings").fetchall()}
        total = reqs["core_credits"] + reqs["elective_courses"] + reqs["free_credits"]
        return (f"Graduation requires {reqs['core_credits']} core credits, "
                f"{reqs['elective_courses']} elective courses, and {reqs['free_credits']} free credits "
                f"({total} total hours).")

    # block / schedule questions (from the block data)
    if "block" in ql:
        lvl = _re.search(r"level\s*(\d+(?:\.\d+)?)", ql)
        bno = _re.search(r"block\s*(\d+)", ql)
        if lvl and bno:
            lv, bn = lvl.group(1), bno.group(1)
            # specific sections with reference numbers (CRNs)
            for key, secs in CHALLENGE_BLOCKS.items():
                kl = key.lower()
                if f"level {lv}" in kl and f"block {bn}" in kl:
                    return f"{key} — sections: " + "; ".join(secs) + "."
            # otherwise the block's course set
            for grp, blocks in BLOCKS.items():
                if lv.split(".")[0] in grp:
                    for b in blocks:
                        if f"block {bn}" in b["name"].lower():
                            return f"{grp} — {b['name']}: " + ", ".join(b["courses"]) + "."
            return f"I don't have a Level {lv} Block {bn} in the plan. Try another block."
        if lvl:
            lv = lvl.group(1)
            for grp, blocks in BLOCKS.items():
                if lv.split(".")[0] in grp:
                    names = "; ".join(b["name"] for b in blocks)
                    return f"{grp} blocks: {names}. Ask for one, e.g. 'level {lv} block 1'."
        return "Available block groups: " + "; ".join(BLOCKS.keys()) + ". Ask e.g. 'level 7 block 1'."

    # level query: "level 7"
    lm = _re.search(r"level\s*(\d{1,2})", ql)
    if lm:
        lvl = int(lm.group(1))
        rows = db.execute("SELECT code,name,credits FROM courses WHERE recommended_level=? ORDER BY code",
                          (lvl,)).fetchall()
        if rows:
            lst = "; ".join(f"{r['code']} {r['name']} ({r['credits']}cr)" for r in rows)
            return f"Level {lvl} courses: {lst}."
        return f"The plan has no courses listed at level {lvl}."

    # keyword search over course names — WHOLE-WORD matches only (no loose guessing)
    kws = [w for w in _re.findall(r"[a-z]+", ql) if w not in _BOT_STOP and len(w) > 2]
    if kws:
        rows = db.execute("SELECT code,name,credits,recommended_level,prereq FROM courses").fetchall()
        hits = [r for r in rows
                if any(_re.search(r"\b" + _re.escape(k) + r"\b", r["name"].lower()) for k in kws)]
        if hits:
            parts = [f"{r['code']} — {r['name']} ({r['credits']}cr, level {r['recommended_level']}"
                     + (f", prereq {r['prereq']}" if r["prereq"] else "") + ")" for r in hits[:6]]
            return "From the CS plan: " + "; ".join(parts) + "."

    # No confident match — do NOT guess.
    return ("I don't have a confident answer for that in the CS plan, so I won't guess. "
            "I can answer plan facts only — a course code (e.g., CPCS-204), a level "
            "(e.g., level 7), a block (e.g., level 7 block 1), or graduation requirements.")


@app.route("/bot/ask", methods=["POST"])
def bot_ask():
    if not current_user():
        abort(403)
    data = request.get_json(silent=True) or {}
    return {"answer": bot_answer(data.get("question", ""))}


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        u = get_db().execute("SELECT * FROM users WHERE email=?", (email,)).fetchone()
        if u and check_password_hash(u["password_hash"], password):
            session["uid"] = u["id"]
            return redirect(url_for(f"{u['role']}_home"))
        flash("Incorrect email or password", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ------------------------- الطالبة -------------------------
@app.route("/student")
@login_required("student")
def student_home():
    db = get_db()
    st = get_student(session["uid"])
    advisor = db.execute("SELECT name FROM users WHERE id=?", (st["advisor_id"],)).fetchone()
    sem = db.execute("SELECT * FROM semesters WHERE is_active=1").fetchone()
    wishes = db.execute(
        """SELECT w.status, c.code, c.name, c.credits, c.type
           FROM wishes w JOIN courses c ON c.id=w.course_id
           WHERE w.student_id=? AND w.semester_id=?
           ORDER BY c.code""",
        (st["id"], sem["id"] if sem else 0),
    ).fetchall()
    return render_template("student_home.html", st=st, advisor=advisor, sem=sem, wishes=wishes)


@app.route("/student/wishes", methods=["GET", "POST"])
@login_required("student")
def student_wishes():
    db = get_db()
    st = get_student(session["uid"])
    sem = db.execute("SELECT * FROM semesters WHERE is_active=1").fetchone()
    if not sem or not sem["wishes_open"]:
        flash("Course preference collection is currently closed", "error")
        return render_template("student_wishes.html", st=st, sem=sem, courses=[], chosen=set())

    if request.method == "POST":
        chosen_ids = set(request.form.getlist("course"))
        # الرغبات المعتمدة/المرفوضة لا تُلمس، نُحدّث فقط المعلّقة
        existing = db.execute(
            "SELECT course_id, status FROM wishes WHERE student_id=? AND semester_id=?",
            (st["id"], sem["id"]),
        ).fetchall()
        locked = {str(r["course_id"]) for r in existing if r["status"] != "pending"}
        # احذف المعلّقة غير المختارة
        db.execute(
            "DELETE FROM wishes WHERE student_id=? AND semester_id=? AND status='pending'",
            (st["id"], sem["id"]),
        )
        for cid in chosen_ids:
            if cid in locked:
                continue
            db.execute(
                "INSERT OR IGNORE INTO wishes(student_id,course_id,semester_id,status) VALUES(?,?,?,'pending')",
                (st["id"], cid, sem["id"]),
            )
        db.commit()
        flash("Your preferences were saved and sent to your advisor for approval", "success")
        return redirect(url_for("student_wishes"))

    courses = db.execute("SELECT * FROM courses WHERE is_active=1 ORDER BY type, code").fetchall()
    rows = db.execute(
        "SELECT course_id, status FROM wishes WHERE student_id=? AND semester_id=?",
        (st["id"], sem["id"]),
    ).fetchall()
    chosen = {str(r["course_id"]) for r in rows}
    status_by_course = {r["course_id"]: r["status"] for r in rows}
    return render_template("student_wishes.html", st=st, sem=sem, courses=courses,
                           chosen=chosen, status_by_course=status_by_course)


@app.route("/student/profile")
@login_required("student")
def student_profile():
    db = get_db()
    st = get_student(session["uid"])
    advisor = db.execute("SELECT name, email FROM users WHERE id=?", (st["advisor_id"],)).fetchone()
    completed = db.execute(
        "SELECT type, COUNT(*) n, SUM(credits) cr FROM completed_courses WHERE student_id=? GROUP BY type",
        (st["id"],),
    ).fetchall()
    done = {r["type"]: {"n": r["n"], "cr": r["cr"]} for r in completed}
    return render_template("student_profile.html", st=st, advisor=advisor,
                           user=current_user(), done=done)


@app.route("/student/task/<key>")
@login_required("student")
def student_task(key):
    task = TASK_BY_KEY.get(key)
    if not task:
        abort(404)
    return render_template("student_task.html", task=task)


def latest_request(student_id, rtype):
    return get_db().execute(
        "SELECT * FROM requests WHERE student_id=? AND type=? ORDER BY id DESC LIMIT 1",
        (student_id, rtype),
    ).fetchone()


@app.route("/student/senior-project", methods=["GET", "POST"])
@login_required("student")
def student_senior_project():
    db = get_db()
    st = get_student(session["uid"])
    existing = latest_request(st["id"], "senior_project")
    if request.method == "POST":
        passed = set(request.form.getlist("passed"))
        eligible = all(code in passed for code, _ in SENIOR_PROJECT_PREREQS)
        details = "Completed courses: " + (", ".join(sorted(passed)) or "None")
        db.execute(
            "INSERT INTO requests(student_id,type,title,details,eligible,status) VALUES(?,?,?,?,?,'pending')",
            (st["id"], "senior_project", "Graduation Project Eligibility (CPCS-498)", details, 1 if eligible else 0),
        )
        db.commit()
        flash("Graduation-project eligibility request sent to your advisor", "success")
        return redirect(url_for("student_senior_project"))
    return render_template("student_senior_project.html", st=st,
                           prereqs=SENIOR_PROJECT_PREREQS, existing=existing)


@app.route("/student/senior-project-team", methods=["GET", "POST"])
@login_required("student")
def student_sp_team():
    db = get_db()
    st = get_student(session["uid"])
    user = current_user()
    advisor = db.execute("SELECT name FROM users WHERE id=?", (st["advisor_id"],)).fetchone()
    existing = latest_request(st["id"], "sp_team")
    team = db.execute("SELECT * FROM sp_teams WHERE student_id=? ORDER BY id DESC LIMIT 1",
                      (st["id"],)).fetchone()
    sp = latest_request(st["id"], "senior_project")  # eligibility status, for a hint
    if request.method == "POST":
        f = request.form
        m1, m1id, m1e = f.get("m1_name", "").strip(), f.get("m1_id", "").strip(), f.get("m1_email", "").strip()
        m2, m2id, m2e = f.get("m2_name", "").strip(), f.get("m2_id", "").strip(), f.get("m2_email", "").strip()
        m3, m3id, m3e = f.get("m3_name", "").strip(), f.get("m3_id", "").strip(), f.get("m3_email", "").strip()
        phone = f.get("phone", "").strip()
        sup_email = f.get("supervisor_email", "").strip()
        sup_ok = 1 if f.get("supervisor_approved") else 0
        comments = f.get("comments", "").strip()
        if not (m1 and m2 and sup_email):
            flash("Please provide at least two members and the project supervisor's email", "error")
            return redirect(url_for("student_sp_team"))
        members = f"1) {m1} ({m1id}); 2) {m2} ({m2id})" + (f"; 3) {m3} ({m3id})" if m3 else "")
        details = (f"Team — {members}. Supervisor: {sup_email} "
                   + ("(approval confirmed)" if sup_ok else "(supervisor approval pending)")
                   + (f". Comments: {comments}" if comments else ""))
        cur = db.execute(
            "INSERT INTO requests(student_id,type,title,details,status) VALUES(?,?,?,?,'pending')",
            (st["id"], "sp_team", "Senior Project Team", details))
        rid = cur.lastrowid
        db.execute(
            """INSERT INTO sp_teams(request_id,student_id,m1_name,m1_id,m1_email,m2_name,m2_id,m2_email,
               m3_name,m3_id,m3_email,phone,supervisor_email,supervisor_approved,comments)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (rid, st["id"], m1, m1id, m1e, m2, m2id, m2e, m3, m3id, m3e, phone, sup_email, sup_ok, comments))
        db.commit()
        flash("Your senior-project team was submitted for approval", "success")
        return redirect(url_for("student_sp_team"))
    return render_template("student_sp_team.html", st=st, user=user, advisor=advisor,
                           existing=existing, team=team, sp=sp)


@app.route("/student/track-certificate", methods=["GET", "POST"])
@login_required("student")
def student_track_cert():
    db = get_db()
    st = get_student(session["uid"])
    existing = latest_request(st["id"], "track_cert")
    selected = TRACK_BY_KEY.get(request.args.get("track"))
    if request.method == "POST":
        track = TRACK_BY_KEY.get(request.form.get("track"))
        courses = request.form.getlist("course")
        if not track or len(courses) < 2:
            flash("Select a track and at least two courses", "error")
            return redirect(url_for("student_track_cert", track=request.form.get("track")))
        details = f"Track: {track['name']} — Courses: " + ", ".join(courses)
        db.execute(
            "INSERT INTO requests(student_id,type,title,details,status) VALUES(?,?,?,?,'pending')",
            (st["id"], "track_cert", "Track Certificate Request", details),
        )
        # Track is determined by the student's elective/track courses:
        # declaring a track certificate sets the student's track.
        db.execute("UPDATE students SET track=? WHERE id=?", (track["name"], st["id"]))
        db.commit()
        flash("Track-certificate request sent to your advisor", "success")
        return redirect(url_for("student_track_cert"))
    return render_template("student_track_cert.html", st=st, tracks=TRACKS,
                           selected=selected, existing=existing)


@app.route("/student/blocks", methods=["GET", "POST"])
@login_required("student")
def student_blocks():
    db = get_db()
    st = get_student(session["uid"])
    existing = latest_request(st["id"], "block")
    group = request.args.get("group") or blocks_group_for_level(st["level"])
    blocks = BLOCKS.get(group, [])
    if request.method == "POST":
        chosen = request.form.get("block")
        grp = request.form.get("group")
        block = next((b for b in BLOCKS.get(grp, []) if b["name"] == chosen), None)
        if not block:
            flash("Please choose a valid block", "error")
            return redirect(url_for("student_blocks", group=grp))
        details = f"{grp} — {block['name']}: " + ", ".join(block["courses"])
        db.execute(
            "INSERT INTO requests(student_id,type,title,details,status) VALUES(?,?,?,?,'pending')",
            (st["id"], "block", "Block Selection", details),
        )
        db.commit()
        flash("Your block choice was sent to your advisor for approval", "success")
        return redirect(url_for("student_blocks"))
    return render_template("student_blocks.html", st=st, group=group,
                           groups=list(BLOCKS.keys()), blocks=blocks, existing=existing)


@app.route("/student/schedule-issues", methods=["GET", "POST"])
@login_required("student")
def student_schedule_issues():
    db = get_db()
    st = get_student(session["uid"])
    existing = latest_request(st["id"], "schedule_issue")
    if request.method == "POST":
        ptype = request.form.get("problem_type", "")
        block = request.form.get("block", "")
        sections = request.form.getlist("section")
        note = request.form.get("note", "").strip()
        details = f"Problem type: {ptype}"
        if block:
            details += f" | Block: {block}"
        if sections:
            details += " | Affected sections: " + ", ".join(sections)
        if note:
            details += f" | Note: {note}"
        db.execute(
            "INSERT INTO requests(student_id,type,title,details,status) VALUES(?,?,?,?,'pending')",
            (st["id"], "schedule_issue", "Schedule registration issue", details),
        )
        db.commit()
        flash("Your issue was sent to your advisor", "success")
        return redirect(url_for("student_schedule_issues"))
    sel_block = request.args.get("block")
    sections = CHALLENGE_BLOCKS.get(sel_block, [])
    return render_template("student_schedule_issues.html", st=st, existing=existing,
                           problem_types=CHALLENGE_TYPES, blocks=list(CHALLENGE_BLOCKS.keys()),
                           sel_block=sel_block, sections=sections)


@app.route("/student/override", methods=["GET", "POST"])
@login_required("student")
def student_override():
    db = get_db()
    st = get_student(session["uid"])
    existing = latest_request(st["id"], "override")
    if request.method == "POST":
        subject = request.form.get("subject", "").strip()
        reason = request.form.get("reason", "").strip()
        if not reason:
            flash("Please describe the reason/details of the waiver request", "error")
            return redirect(url_for("student_override"))
        details = (f"Regarding: {subject} | " if subject else "") + f"Details: {reason}"
        db.execute(
            "INSERT INTO requests(student_id,type,title,details,status) VALUES(?,?,?,?,'pending')",
            (st["id"], "override", "Waiver Request", details),
        )
        db.commit()
        flash("Your waiver request was sent to your advisor to handle", "success")
        return redirect(url_for("student_override"))
    courses = db.execute(
        "SELECT code, name FROM courses WHERE is_active=1 ORDER BY code"
    ).fetchall()
    return render_template("student_override.html", st=st, existing=existing, courses=courses)


@app.route("/student/summer-training", methods=["GET", "POST"])
@login_required("student")
def student_summer_training():
    db = get_db()
    st = get_student(session["uid"])
    existing = latest_request(st["id"], "summer_training")
    # الشرط: إكمال جميع مقررات المستويين السابع والثامن
    courses = db.execute(
        """SELECT code, name, recommended_level FROM courses
           WHERE recommended_level IN (7,8) AND is_active=1 AND code!='CPCS-323'
           ORDER BY recommended_level, code"""
    ).fetchall()
    if request.method == "POST":
        passed = set(request.form.getlist("passed"))
        missing = sorted({c["code"] for c in courses} - passed)
        n = len(missing)
        if n > 2:
            flash("Cannot submit: more than two courses are missing — summer training is not allowed.", "error")
            return redirect(url_for("student_summer_training"))
        if n == 0:
            title = "Summer Training Request (CPCS-323)"
            details = "Completed all level 7 & 8 courses — eligible."
            eligible = 1
        else:  # 1 or 2 courses missing → waiver required
            title = "Summer Training Request — Waiver Required"
            details = (f"Missing {n} course(s): " + ", ".join(missing) +
                       ". Requires an exceptional (waiver) approval from the department.")
            eligible = 0
        db.execute(
            "INSERT INTO requests(student_id,type,title,details,eligible,status) VALUES(?,?,?,?,?,'pending')",
            (st["id"], "summer_training", title, details, eligible),
        )
        db.commit()
        flash("Summer training request sent to your advisor" +
              ("" if eligible else " (waiver required)"), "success")
        return redirect(url_for("student_summer_training"))
    return render_template("student_summer_training.html", st=st, courses=courses,
                           total=len(courses), existing=existing)


@app.route("/student/graduation", methods=["GET", "POST"])
@login_required("student")
def student_graduation():
    db = get_db()
    st = get_student(session["uid"])
    report = graduation_report(st["id"])
    existing = latest_request(st["id"], "grad_confirm")
    if request.method == "POST":
        grad_sem = request.form.get("grad_semester", "").strip()
        remaining = ", ".join(
            f"{it['label']}: {it['remaining']} {it['unit']} remaining"
            for it in report["items"] if not it["met"]
        ) or "No remaining requirements — all complete"
        details = f"Student graduation confirmation. Status: {remaining}"
        db.execute(
            "INSERT INTO requests(student_id,type,title,details,eligible,grad_semester,status) VALUES(?,?,?,?,?,?,'pending')",
            (st["id"], "grad_confirm", "Graduation Confirmation Request", details,
             1 if report["eligible"] else 0, grad_sem),
        )
        db.commit()
        flash("Graduation confirmation request sent to your advisor", "success")
        return redirect(url_for("student_graduation"))
    return render_template("student_graduation.html", st=st, report=report,
                           existing=existing, grad_semesters=GRAD_SEMESTERS)


@app.route("/student/graduation-letter")
@login_required("student")
def student_grad_letter():
    db = get_db()
    st = get_student(session["uid"])
    user = current_user()
    advisor = db.execute("SELECT name FROM users WHERE id=?", (st["advisor_id"],)).fetchone()
    req = latest_request(st["id"], "grad_confirm")
    approved = req if req and req["status"] == "approved" else None
    reqs = {r["key"]: int(r["value"]) for r in db.execute("SELECT key,value FROM settings").fetchall()}
    total = reqs.get("core_credits", 121) + reqs.get("elective_courses", 9) + reqs.get("free_credits", 10)
    return render_template("student_grad_letter.html", st=st, user=user,
                           advisor=advisor, req=req, approved=approved, reqs=reqs, total=total)


def graduation_report(student_id):
    db = get_db()
    reqs = {r["key"]: int(r["value"]) for r in db.execute("SELECT key,value FROM settings").fetchall()}
    rows = db.execute(
        "SELECT type, credits FROM completed_courses WHERE student_id=?", (student_id,)
    ).fetchall()
    core = sum(r["credits"] for r in rows if r["type"] == "core")
    elective_count = sum(1 for r in rows if r["type"] == "elective")
    free = sum(r["credits"] for r in rows if r["type"] == "free")
    items = [
        {"label": "Core Credits", "have": core, "need": reqs.get("core_credits", 121), "unit": "credits"},
        {"label": "Elective Courses", "have": elective_count, "need": reqs.get("elective_courses", 9), "unit": "courses"},
        {"label": "Free Credits", "have": free, "need": reqs.get("free_credits", 10), "unit": "credits"},
    ]
    for it in items:
        it["remaining"] = max(0, it["need"] - it["have"])
        it["met"] = it["have"] >= it["need"]
        it["pct"] = min(100, round(it["have"] / it["need"] * 100)) if it["need"] else 100
    items_all_met = all(it["met"] for it in items)
    return {"items": items, "eligible": items_all_met}


# ------------------------- المرشدة -------------------------
@app.route("/advisor")
@login_required("advisor")
def advisor_home():
    return redirect(url_for("advisor_students"))


@app.route("/advisor/students")
@login_required("advisor")
def advisor_students():
    db = get_db()
    students = db.execute(
        """SELECT s.id, u.name, s.university_id, s.level, s.track,
                  (SELECT COUNT(*) FROM wishes w WHERE w.student_id=s.id AND w.status='pending') AS pending
           FROM students s JOIN users u ON u.id=s.user_id
           WHERE s.advisor_id=? ORDER BY s.level, u.name""",
        (session["uid"],),
    ).fetchall()
    return render_template("advisor_students.html", students=students)


@app.route("/advisor/approvals", methods=["GET", "POST"])
@login_required("advisor")
def advisor_approvals():
    db = get_db()
    if request.method == "POST":
        wish_id = request.form.get("wish_id")
        action = request.form.get("action")
        note = request.form.get("note", "")
        # تأكد أن الرغبة تخص طالبة تحت إشراف هذه المرشدة
        row = db.execute(
            """SELECT w.id FROM wishes w JOIN students s ON s.id=w.student_id
               WHERE w.id=? AND s.advisor_id=?""",
            (wish_id, session["uid"]),
        ).fetchone()
        if row and action in ("approved", "rejected"):
            db.execute("UPDATE wishes SET status=?, advisor_note=? WHERE id=?",
                       (action, note, wish_id))
            db.commit()
            flash("Preference status updated", "success")
        return redirect(url_for("advisor_approvals"))

    pending = db.execute(
        """SELECT w.id, u.name AS student, s.university_id, s.level,
                  c.code, c.name AS course, c.type, c.credits
           FROM wishes w
           JOIN students s ON s.id=w.student_id
           JOIN users u ON u.id=s.user_id
           JOIN courses c ON c.id=w.course_id
           WHERE s.advisor_id=? AND w.status='pending'
           ORDER BY u.name, c.code""",
        (session["uid"],),
    ).fetchall()
    return render_template("advisor_approvals.html", pending=pending)


@app.route("/advisor/requests", methods=["GET", "POST"])
@login_required("advisor")
def advisor_requests():
    db = get_db()
    if request.method == "POST":
        rid = request.form.get("request_id")
        action = request.form.get("action")
        note = request.form.get("note", "")
        row = db.execute(
            """SELECT r.id FROM requests r JOIN students s ON s.id=r.student_id
               WHERE r.id=? AND s.advisor_id=?""",
            (rid, session["uid"]),
        ).fetchone()
        if row and action in ("approved", "rejected"):
            db.execute("UPDATE requests SET status=?, advisor_note=? WHERE id=?",
                       (action, note, rid))
            db.commit()
            flash("Request status updated", "success")
        return redirect(url_for("advisor_requests"))

    reqs = db.execute(
        """SELECT r.*, u.name AS student, s.university_id, s.level
           FROM requests r
           JOIN students s ON s.id=r.student_id
           JOIN users u ON u.id=s.user_id
           WHERE s.advisor_id=?
           ORDER BY CASE r.status WHEN 'pending' THEN 0 ELSE 1 END, r.id DESC""",
        (session["uid"],),
    ).fetchall()
    return render_template("advisor_requests.html", reqs=reqs)


@app.route("/advisor/stats")
@login_required("advisor")
def advisor_stats():
    stats = course_stats()
    levels = sorted({lvl for s in stats for lvl in s["by_level"]})
    return render_template("advisor_stats.html", stats=stats, levels=levels)


@app.route("/advisor/stats/course/<int:course_id>")
@login_required("advisor")
def advisor_stats_course(course_id):
    """تفاصيل الطالبات الراغبات في مقرر معيّن (الضغط على الأرقام)."""
    db = get_db()
    course = db.execute("SELECT * FROM courses WHERE id=?", (course_id,)).fetchone()
    if not course:
        abort(404)
    sem = db.execute("SELECT * FROM semesters WHERE is_active=1").fetchone()
    students = db.execute(
        """SELECT u.name, s.university_id, s.level, s.track, w.status
           FROM wishes w
           JOIN students s ON s.id=w.student_id
           JOIN users u ON u.id=s.user_id
           WHERE w.course_id=? AND w.semester_id=?
           ORDER BY s.level, u.name""",
        (course_id, sem["id"] if sem else 0),
    ).fetchall()
    return render_template("advisor_course_detail.html", course=course, students=students)


def course_stats():
    """إحصائية لكل مقرر: عدد الراغبات + التوزيع حسب المستوى (للفصل الفعّال)."""
    db = get_db()
    sem = db.execute("SELECT * FROM semesters WHERE is_active=1").fetchone()
    if not sem:
        return []
    rows = db.execute(
        """SELECT c.id, c.code, c.name, c.type, s.level, w.status
           FROM wishes w
           JOIN courses c ON c.id=w.course_id
           JOIN students s ON s.id=w.student_id
           WHERE w.semester_id=?""",
        (sem["id"],),
    ).fetchall()
    agg = {}
    for r in rows:
        key = r["id"]
        if key not in agg:
            agg[key] = {"id": r["id"], "code": r["code"], "name": r["name"],
                        "type": r["type"], "total": 0, "approved": 0, "by_level": {}}
        agg[key]["total"] += 1
        if r["status"] == "approved":
            agg[key]["approved"] += 1
        agg[key]["by_level"][r["level"]] = agg[key]["by_level"].get(r["level"], 0) + 1
    return sorted(agg.values(), key=lambda x: (-x["total"], x["code"]))


@app.route("/advisor/stats/export")
@login_required("advisor")
def advisor_stats_export():
    import openpyxl
    stats = course_stats()
    levels = sorted({lvl for s in stats for lvl in s["by_level"]})
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Course Statistics"
    header = ["Code", "Course", "Type", "Total Interested", "Approved"] + \
             [f"Level {l}" for l in levels]
    ws.append(header)
    for s in stats:
        ws.append([s["code"], s["name"], TYPE_AR.get(s["type"], s["type"]),
                   s["total"], s["approved"]] + [s["by_level"].get(l, 0) for l in levels])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, attachment_filename="compass_stats.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ------------------------- المسؤول -------------------------
@app.route("/admin")
@login_required("admin")
def admin_home():
    db = get_db()
    counts = {
        "students": db.execute("SELECT COUNT(*) FROM students").fetchone()[0],
        "advisors": db.execute("SELECT COUNT(*) FROM users WHERE role='advisor'").fetchone()[0],
        "courses": db.execute("SELECT COUNT(*) FROM courses").fetchone()[0],
        "wishes": db.execute("SELECT COUNT(*) FROM wishes").fetchone()[0],
    }
    return render_template("admin_home.html", counts=counts)


@app.route("/admin/courses", methods=["GET", "POST"])
@login_required("admin")
def admin_courses():
    db = get_db()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            try:
                db.execute(
                    "INSERT INTO courses(code,name,type,credits,parity,recommended_level,prereq) VALUES(?,?,?,?,?,?,?)",
                    (request.form["code"].strip().upper(), request.form["name"].strip(),
                     request.form["type"], int(request.form["credits"]),
                     request.form["parity"], int(request.form.get("level") or 0) or None,
                     request.form.get("prereq", "").strip()),
                )
                db.commit()
                flash("Course added", "success")
            except Exception:
                flash("Duplicate course code or invalid data", "error")
        elif action == "toggle":
            db.execute("UPDATE courses SET is_active = 1 - is_active WHERE id=?",
                       (request.form["id"],))
            db.commit()
        return redirect(url_for("admin_courses"))
    courses = db.execute("SELECT * FROM courses ORDER BY type, code").fetchall()
    return render_template("admin_courses.html", courses=courses)


@app.route("/admin/semesters", methods=["GET", "POST"])
@login_required("admin")
def admin_semesters():
    db = get_db()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            db.execute("INSERT INTO semesters(name) VALUES(?)", (request.form["name"].strip(),))
            db.commit()
            flash("Semester added", "success")
        elif action == "activate":
            db.execute("UPDATE semesters SET is_active=0")
            db.execute("UPDATE semesters SET is_active=1 WHERE id=?", (request.form["id"],))
            db.commit()
        elif action == "toggle_wishes":
            db.execute("UPDATE semesters SET wishes_open = 1 - wishes_open WHERE id=?",
                       (request.form["id"],))
            db.commit()
        return redirect(url_for("admin_semesters"))
    semesters = db.execute("SELECT * FROM semesters ORDER BY id DESC").fetchall()
    return render_template("admin_semesters.html", semesters=semesters)


@app.route("/admin/users")
@login_required("admin")
def admin_users():
    db = get_db()
    users = db.execute(
        """SELECT u.id, u.name, u.email, u.role, s.university_id, s.level, s.track,
                  a.name AS advisor
           FROM users u
           LEFT JOIN students s ON s.user_id=u.id
           LEFT JOIN users a ON a.id=s.advisor_id
           ORDER BY u.role, u.name""",
    ).fetchall()
    return render_template("admin_users.html", users=users)


def graduates_rows():
    """صف لكل طالبة قدّمت طلب تأكيد تخرج (أحدث طلب لكل طالبة)."""
    db = get_db()
    return db.execute(
        """SELECT u.name AS student, s.university_id, s.level, s.track,
                  a.name AS advisor, r.status, r.eligible, r.grad_semester,
                  r.advisor_note, r.created_at
           FROM requests r
           JOIN students s ON s.id=r.student_id
           JOIN users u ON u.id=s.user_id
           LEFT JOIN users a ON a.id=s.advisor_id
           WHERE r.type='grad_confirm'
             AND r.id=(SELECT MAX(id) FROM requests r2
                       WHERE r2.student_id=r.student_id AND r2.type='grad_confirm')
           ORDER BY CASE r.status WHEN 'approved' THEN 0 WHEN 'pending' THEN 1 ELSE 2 END,
                    u.name""",
    ).fetchall()


@app.route("/admin/graduates")
@login_required("admin")
def admin_graduates():
    rows = graduates_rows()
    summary = {
        "total": len(rows),
        "approved": sum(1 for r in rows if r["status"] == "approved"),
        "pending": sum(1 for r in rows if r["status"] == "pending"),
        "eligible": sum(1 for r in rows if r["eligible"]),
    }
    return render_template("admin_graduates.html", rows=rows, summary=summary)


@app.route("/admin/graduates/export")
@login_required("admin")
def admin_graduates_export():
    import openpyxl
    rows = graduates_rows()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Graduates"
    ws.append(["Student", "University ID", "Level", "Track", "Advisor",
               "Grad. Semester", "Advisor Approval", "Requirements Met", "Request Date", "Note"])
    for r in rows:
        ws.append([
            r["student"], r["university_id"], r["level"], r["track"], r["advisor"] or "",
            r["grad_semester"] or "", STATUS_AR.get(r["status"], r["status"]),
            "Met" if r["eligible"] else "Not met", r["created_at"] or "", r["advisor_note"] or "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, attachment_filename="compass_graduates.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def teams_rows():
    """One row per submitted senior-project team, with student + advisor + approval status."""
    return get_db().execute(
        """SELECT t.*, u.name AS student, s.university_id, a.name AS advisor,
                  r.status AS req_status, r.advisor_note
           FROM sp_teams t
           JOIN students s ON s.id=t.student_id
           JOIN users u ON u.id=s.user_id
           LEFT JOIN users a ON a.id=s.advisor_id
           LEFT JOIN requests r ON r.id=t.request_id
           ORDER BY t.id DESC""").fetchall()


@app.route("/admin/teams")
@login_required("admin")
def admin_teams():
    return render_template("admin_teams.html", rows=teams_rows())


@app.route("/admin/teams/export")
@login_required("admin")
def admin_teams_export():
    import openpyxl
    rows = teams_rows()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Senior Project Teams"
    ws.append(["Submitted by", "University ID", "Advisor", "Member 1", "M1 ID", "M1 Email",
               "Member 2", "M2 ID", "M2 Email", "Member 3", "M3 ID", "M3 Email", "Phone",
               "Supervisor Email", "Supervisor Approved", "Advisor Approval", "Comments"])
    for r in rows:
        ws.append([
            r["student"], r["university_id"], r["advisor"] or "",
            r["m1_name"], r["m1_id"], r["m1_email"],
            r["m2_name"], r["m2_id"], r["m2_email"],
            r["m3_name"] or "", r["m3_id"] or "", r["m3_email"] or "",
            r["phone"] or "", r["supervisor_email"] or "",
            "Yes" if r["supervisor_approved"] else "No",
            STATUS_AR.get(r["req_status"], r["req_status"] or ""), r["comments"] or "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, attachment_filename="compass_sp_teams.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=5000, debug=True)
